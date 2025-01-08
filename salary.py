import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import random

# Employee data for basic pay lookup
employee_basic_pay = {
    "prakash ghodeswar": 10000,
    "sonu sharma": 9500,
    "shriram ninave": 10000,
    "anand kathwate": 20000,
    "kamlesh goyal": 20000,
    "umesh borkar": 12500,
    "manav sur": 20000,
    "ankit khobragade": 10000,
    "anand motghare": 20000,
    "niraj kidey": 10000,
    "prakash nimkar": 12500,
    "anand kumbhare": 12500,
    "chandrakant panchariya": 20000,
    "vijay ulabhaje": 10000,
    "rudesh gedam": 20000,
    "ankush dhawankar": 10000,
    "ramchanda belekar": 10000,
    "ravi pimpalkar": 10000,
    "mahesh salpekar": 10000,
    "kirti chandak": 10000,
    "amit mundada": 20000,
    "govind bhutada": 10000,
    "akash haygune": 9500,
    "sunil katekar": 10000,
    "shrikant paturkar": 12500
}

# List of employees who should not receive PF or have reduced city allowance or gratuity
employees_no_pf = ["shriram ninave", "sunil katekar", "mahesh salpekar", "kirti chandak", "ramchandra belekar", "ravi pimpalkar", "ankush dhawankar"]
employees_reduced_city_allowance = ["prakash ghodeswar", "sonu sharma", "amit mundada"]
employees_no_gratuity = ["sunil katekar", "ramchanda belekar", "mahesh salpekar", "ankush dhawankar", "shriram ninave"]
employees_no_conveyance = ["kirti chandak", "sonu sharma", "sunil katekar", "ramchanda belekar", "mahesh salpekar", "ankush dhawankar"]

# Function to validate numeric inputs
def is_numeric(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

# Function to calculate salary and deductions
def calculate_salary(emp_name, company_name, month, post, days_worked, advance_deduction, tds_deduction):
    try:
        emp_name = emp_name.strip().lower()  # Ensure no leading/trailing spaces and convert to lowercase
        basic_salary = employee_basic_pay.get(emp_name, 0)

        if basic_salary == 0:
            st.error(f"Employee '{emp_name}' not found! Please check the name entered.")
            return

        # Adjusted salary components
        total_days_in_month = 26
        adjusted_basic_salary = (basic_salary / total_days_in_month) * days_worked
        hra = round(adjusted_basic_salary * 0.40, 2)
        conveyance = round(adjusted_basic_salary * 0.20, 2)

        # Deductions and bonuses
        epf = round(adjusted_basic_salary * 0.13, 2) if emp_name not in employees_no_pf else 0
        provident_fund = round(adjusted_basic_salary * 0.12, 2) if emp_name not in employees_no_pf else 0
        gratuity = round(adjusted_basic_salary * 0.0481, 2)
        
        gross_salary = adjusted_basic_salary + hra + conveyance
        leave_deduction = round(gross_salary * 0.09705, 2)
        if leave_deduction > 18633:
            leave_deduction = 0

        bonus = round(adjusted_basic_salary * 0.0833, 2)
        professional_tax = 200 if adjusted_basic_salary > 7500 else 175

        city_allowance = round(adjusted_basic_salary * random.uniform(1.212, 1.212), 2)
        esic = 0  # Assuming 0 ESIC as this is for simplification

        total_emoluments = gross_salary + city_allowance + bonus + leave_deduction + epf + professional_tax + gratuity
        total_emoluments_12_months = total_emoluments * 12

        total_deductions = epf + provident_fund + gratuity + leave_deduction + professional_tax + esic
        take_home_salary = gross_salary + city_allowance - total_deductions + provident_fund + bonus - professional_tax + leave_deduction

        final_take_home_salary = take_home_salary - advance_deduction - tds_deduction

        save_to_excel(company_name, month, post, emp_name, adjusted_basic_salary, hra, conveyance, city_allowance, epf, provident_fund, gratuity, leave_deduction, bonus, professional_tax, esic, final_take_home_salary, total_emoluments_12_months, advance_deduction, tds_deduction)

        st.write(f"Calculated Salary for {emp_name.capitalize()}:")
        st.write(f"Take Home Salary: ₹{final_take_home_salary:.2f}")
        st.write(f"Total Emoluments (TEC): ₹{total_emoluments_12_months:.2f}")

    except Exception as e:
        st.error(f"Calculation error: {str(e)}")

# Function to save data to Excel
def save_to_excel(company_name, month, post, emp_name, basic_rate, hra, conveyance, city_allowance, epf, provident_fund, gratuity, leave_deduction, bonus, professional_tax, esic, take_home_salary, total_emoluments_12_months, advance_deduction, tds_deduction):
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Details"

    # Define header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Add headers
    headers = ["Field", "Value"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add employee data
    data = [
        ("Company Name", company_name),
        ("Month", month),
        ("Post", post),
        ("Employee Name", emp_name),
        ("Basic Rate (Calculated)", basic_rate),
        ("HRA (40% of Basic Pay)", hra),
        ("Conveyance (20% of Basic Pay)", conveyance),
        ("City Allowance", city_allowance),
        ("EPF Contribution", epf),
        ("Provident Fund Contribution", provident_fund),
        ("Gratuity", gratuity),
        ("Leave Deduction", leave_deduction),
        ("Bonus", bonus),
        ("Professional Tax", professional_tax),
        ("ESIC (8.05% of Basic Pay)", esic),
        ("Advance Deduction", advance_deduction),
        ("TDS Deduction", tds_deduction),
        ("Take Home Salary", take_home_salary),
        ("Total Emoluments (TEC)", total_emoluments_12_months)
    ]

    for row_num, (field, value) in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=field)
        ws.cell(row=row_num, column=2, value=value)

    file_name = f"{emp_name}_Salary_Details.xlsx"
    wb.save(file_name)
    st.success(f"File saved as {file_name}")

# Streamlit UI
st.title("Employee Salary Calculation")
company_name = st.text_input("Company Name")
month = st.text_input("Month")
post = st.text_input("Post")
emp_name = st.text_input("Employee Name")
days_worked = st.number_input("Days Worked", min_value=1, max_value=31)
advance_deduction = st.number_input("Advance Deduction", min_value=0.0)
tds_deduction = st.number_input("TDS Deduction", min_value=0.0)

if st.button("Calculate Salary"):
    calculate_salary(emp_name, company_name, month, post, days_worked, advance_deduction, tds_deduction)
